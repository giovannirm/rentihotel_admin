# coding:utf-8
"""
Utilidades Varias
=================
"""

import openpyxl
import random
from django.http import HttpResponse


class Numbers:
    """
    Clase de utilidades para trabajar con datos de tipo numérico.
    """

    @staticmethod
    def is_number(s):
        """
        Comprueba que un dato es del tipo numérico o contiene números.

        :param s: Dato que será comprobado.
        :return: ``True`` si ``s`` es numérico, ``False`` de otra forma.

        Uso:

        .. code:: python

            Numbers.is_number('123')  # True
            Numbers.is_number('uno')  # False
            Numbers.is_number(123)    # True
        """

        try:
            float(s)
            return True
        except ValueError:
            pass

        try:
            import unicodedata
            unicodedata.numeric(s)
            return True
        except (TypeError, ValueError):
            pass

        return False


class RandomUtil:
    """
    Clase de utilidades para trabajar con números aleatorios.
    """

    @staticmethod
    def get_random_probability(option_probability_dict=None):
        """
        TODO: Llenar

        :param option_probability_dict: To be implemented
        :return: None
        """

        random_number = random.random()
        probabilities_sum = 0.0
        prize_id = 0

        for key in option_probability_dict:

            probability = option_probability_dict[key] / 100.0

            probability_max = probability + probabilities_sum

            if random_number <= probability_max:
                prize_id = key
                break

            probabilities_sum += probability

        return prize_id


class ExcelResponse(HttpResponse):
    """
    ExcelResponse(HttpResponse)

    Clase de generación de archivos .xlsx

    Ejemplo:

    .. code:: python

        class Report(View):
            def get(self, request, *args, **kwargs):
                return ExcelResponse(
                    [('A1', 'B1', 'C1'), ('A2', 'B2', 'C2')],
                    'report'
                )

    .. note::

        Requiere la instalación del módulo ``openpyxl``.

    :param data: Celdas que contendrá el archivo.
    :param output_name: Nombre del archivo a generar. La extensión se agrega automáticamente.
    :param headers: Lista de cabeceras para el archivo. Opcional.
    """

    def __init__(self, data, output_name='excel_data', headers=None):

        # Make sure we've got the right type of data to work with
        valid_data = False
        if hasattr(data, '__getitem__') and not type(data) == dict:
            if isinstance(data[0], dict):
                if headers is None:
                    headers = data[0].keys()
                data = [[row[col] for col in headers] for row in data]
                data.insert(0, headers)
            if hasattr(data[0], '__getitem__'):
                valid_data = True
        elif type(data) == dict:
            for value in data.values():
                if isinstance(value[0], dict):
                    if headers is None:
                        headers = value[0].keys()
                    value = [[row[col] for col in headers] for row in value]
                    value.insert(0, headers)
                if hasattr(value[0], '__getitem__'):
                    valid_data = True

        assert valid_data is True, \
            "ExcelResponse requires a sequence of sequences or a dictionary " \
            "with sequences as values"

        import StringIO
        output = StringIO.StringIO()

        book = openpyxl.Workbook()

        if type(data) != dict:
            sheet = book.active

            for rowx, row in enumerate(data):
                for colx, value in enumerate(row):
                    cell = sheet.cell(row=rowx+1, column=colx+1)
                    cell.value = value
        else:
            for key, val in data.items():
                del book.worksheets[0]
                sheet = book.create_sheet(title=key)

                for rowx, row in enumerate(val):
                    for colx, value in enumerate(row):
                        cell = sheet.cell(row=rowx+1, column=colx+1)
                        cell.value = value

        book.save(output)

        mimetype = \
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        file_ext = 'xlsx'

        output.seek(0)
        super(ExcelResponse, self).__init__(content=output.getvalue(),
                                            content_type=mimetype)
        self['Content-Disposition'] = 'attachment;filename="%s.%s"' % \
            (output_name.replace('"', '\"'), file_ext)
